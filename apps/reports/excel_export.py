from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


class ReportXLSXRenderer:
    @staticmethod
    def _style_header(cell):
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="1F2937",
            end_color="1F2937",
            fill_type="solid",
        )

    @staticmethod
    def _autofit_columns(worksheet, max_col: int):
        for col_index in range(1, max_col + 1):
            max_length = 0
            col_letter = worksheet.cell(row=1, column=col_index).column_letter
            for cell in worksheet[col_letter]:
                value = "" if cell.value is None else str(cell.value)
                if len(value) > max_length:
                    max_length = len(value)
            worksheet.column_dimensions[col_letter].width = min(max_length + 2, 40)

    @staticmethod
    def render_member_statement(payload: dict, *, watermark: bool = False) -> bytes:
        workbook = Workbook()

        summary_ws = workbook.active
        summary_ws.title = "Summary"
        summary_ws["A1"] = "Member Statement"
        summary_ws["A1"].font = Font(size=16, bold=True)
        summary_ws["A2"] = f"Member: {payload.get('member_name')}"
        summary_ws["A3"] = (
            f"Period: {payload.get('from') or 'N/A'} to {payload.get('to') or 'N/A'}"
        )
        if watermark:
            summary_ws["A4"] = "AUDITOR COPY"
            summary_ws["A4"].font = Font(bold=True, color="B91C1C")

        totals = payload.get("totals", {})
        summary_ws["A6"] = "Metric"
        summary_ws["B6"] = "Value"
        ReportXLSXRenderer._style_header(summary_ws["A6"])
        ReportXLSXRenderer._style_header(summary_ws["B6"])

        rows = [
            ("Contributions", totals.get("contributions", "0.00")),
            ("Loan Disbursements", totals.get("loan_disbursements", "0.00")),
            ("Repayments", totals.get("repayments", "0.00")),
            ("Penalties Debited", totals.get("penalties_debited", "0.00")),
            ("Penalties Credited", totals.get("penalties_credited", "0.00")),
            ("Closing Balance", totals.get("closing_balance", "0.00")),
        ]
        for idx, (metric, value) in enumerate(rows, start=7):
            summary_ws.cell(row=idx, column=1, value=metric)
            summary_ws.cell(row=idx, column=2, value=value)

        ledger_ws = workbook.create_sheet("Ledger")
        headers = [
            "Date",
            "Entry Type",
            "Direction",
            "Amount",
            "Running Balance",
            "Narration",
        ]
        for col, header in enumerate(headers, start=1):
            ledger_ws.cell(row=1, column=col, value=header)
            ReportXLSXRenderer._style_header(ledger_ws.cell(row=1, column=col))

        for row_index, line in enumerate(payload.get("ledger", []), start=2):
            ledger_ws.cell(row=row_index, column=1, value=line.get("date"))
            ledger_ws.cell(row=row_index, column=2, value=line.get("entry_type"))
            ledger_ws.cell(row=row_index, column=3, value=line.get("direction"))
            ledger_ws.cell(row=row_index, column=4, value=line.get("amount"))
            ledger_ws.cell(row=row_index, column=5, value=line.get("running_balance"))
            ledger_ws.cell(row=row_index, column=6, value=line.get("narration"))

        ReportXLSXRenderer._autofit_columns(summary_ws, 2)
        ReportXLSXRenderer._autofit_columns(ledger_ws, 6)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.read()

    @staticmethod
    def render_chama_summary(payload: dict, *, watermark: bool = False) -> bytes:
        workbook = Workbook()

        summary_ws = workbook.active
        summary_ws.title = "Summary"
        summary_ws["A1"] = "Chama Monthly Summary"
        summary_ws["A1"].font = Font(size=16, bold=True)
        summary_ws["A2"] = f"Chama: {payload.get('chama_name')}"
        summary_ws["A3"] = f"Period: {payload.get('month')}/{payload.get('year')}"
        if watermark:
            summary_ws["A4"] = "AUDITOR COPY"
            summary_ws["A4"].font = Font(bold=True, color="B91C1C")

        summary_ws["A6"] = "Metric"
        summary_ws["B6"] = "Value"
        ReportXLSXRenderer._style_header(summary_ws["A6"])
        ReportXLSXRenderer._style_header(summary_ws["B6"])

        totals = payload.get("totals", {})
        cashflow = payload.get("cashflow", {})
        rows = [
            ("Contributions", totals.get("contributions", "0.00")),
            ("Repayments", totals.get("repayments", "0.00")),
            ("Penalties Issued", totals.get("penalties_issued", "0.00")),
            ("Penalties Collected", totals.get("penalties_collected", "0.00")),
            ("Loans Out", totals.get("loans_out", "0.00")),
            ("Cashflow Credits", cashflow.get("credits", "0.00")),
            ("Cashflow Debits", cashflow.get("debits", "0.00")),
            ("Net Cashflow", cashflow.get("net", "0.00")),
        ]
        for idx, (metric, value) in enumerate(rows, start=7):
            summary_ws.cell(row=idx, column=1, value=metric)
            summary_ws.cell(row=idx, column=2, value=value)

        defaulters_ws = workbook.create_sheet("Defaulters")
        headers = [
            "Member Name",
            "Member Phone",
            "Loan Status",
            "Outstanding Balance",
            "Overdue Installments",
        ]
        for col, header in enumerate(headers, start=1):
            defaulters_ws.cell(row=1, column=col, value=header)
            ReportXLSXRenderer._style_header(defaulters_ws.cell(row=1, column=col))

        for row_index, row in enumerate(payload.get("defaulters", []), start=2):
            defaulters_ws.cell(row=row_index, column=1, value=row.get("member_name"))
            defaulters_ws.cell(row=row_index, column=2, value=row.get("member_phone"))
            defaulters_ws.cell(row=row_index, column=3, value=row.get("status"))
            defaulters_ws.cell(
                row=row_index, column=4, value=row.get("outstanding_balance")
            )
            defaulters_ws.cell(
                row=row_index,
                column=5,
                value=row.get("overdue_installments"),
            )

        ReportXLSXRenderer._autofit_columns(summary_ws, 2)
        ReportXLSXRenderer._autofit_columns(defaulters_ws, 5)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.read()

    @staticmethod
    def render_loan_schedule(payload: dict, *, watermark: bool = False) -> bytes:
        workbook = Workbook()
        ws = workbook.active
        ws.title = "Loan Schedule"
        ws["A1"] = "Loan Repayment Schedule"
        ws["A1"].font = Font(size=16, bold=True)
        loan = payload.get("loan", {})
        ws["A2"] = f"Loan ID: {loan.get('loan_id')}"
        ws["A3"] = f"Member: {loan.get('member_name')}"
        if watermark:
            ws["A4"] = "AUDITOR COPY"
            ws["A4"].font = Font(bold=True, color="B91C1C")

        headers = [
            "Due Date",
            "Expected Amount",
            "Principal",
            "Interest",
            "Penalty",
            "Status",
        ]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=6, column=col, value=header)
            ReportXLSXRenderer._style_header(ws.cell(row=6, column=col))

        for row_index, item in enumerate(payload.get("schedule", []), start=7):
            ws.cell(row=row_index, column=1, value=item.get("due_date"))
            ws.cell(row=row_index, column=2, value=item.get("expected_amount"))
            ws.cell(row=row_index, column=3, value=item.get("expected_principal"))
            ws.cell(row=row_index, column=4, value=item.get("expected_interest"))
            ws.cell(row=row_index, column=5, value=item.get("expected_penalty"))
            ws.cell(row=row_index, column=6, value=item.get("status"))

        ReportXLSXRenderer._autofit_columns(ws, 6)
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.read()

    @staticmethod
    def render_loan_approvals_log(payload: dict, *, watermark: bool = False) -> bytes:
        workbook = Workbook()
        ws = workbook.active
        ws.title = "Loan Approvals"
        ws["A1"] = "Loan Approvals Log"
        ws["A1"].font = Font(size=16, bold=True)
        ws["A2"] = f"Chama: {payload.get('chama_name')}"
        if payload.get("month") and payload.get("year"):
            ws["A3"] = f"Period: {payload.get('month')}/{payload.get('year')}"
        if watermark:
            ws["A4"] = "AUDITOR COPY"
            ws["A4"].font = Font(bold=True, color="B91C1C")

        headers = [
            "Loan ID",
            "Member",
            "Stage",
            "Decision",
            "Actor",
            "Acted At",
            "Note",
        ]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=6, column=col, value=header)
            ReportXLSXRenderer._style_header(ws.cell(row=6, column=col))

        for row_index, row in enumerate(payload.get("rows", []), start=7):
            ws.cell(row=row_index, column=1, value=row.get("loan_id"))
            ws.cell(row=row_index, column=2, value=row.get("member_name"))
            ws.cell(row=row_index, column=3, value=row.get("stage"))
            ws.cell(row=row_index, column=4, value=row.get("decision"))
            ws.cell(row=row_index, column=5, value=row.get("actor_name"))
            ws.cell(row=row_index, column=6, value=row.get("acted_at"))
            ws.cell(row=row_index, column=7, value=row.get("note"))

        ReportXLSXRenderer._autofit_columns(ws, 7)
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.read()
