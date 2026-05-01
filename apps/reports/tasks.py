"""
Celery Tasks for Report Generation

Async report generation with file output.
"""

import io
import json
import logging
from datetime import datetime, timedelta

from celery import shared_task
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.utils import timezone

logger = logging.getLogger(__name__)


def _report_storage_name(file_name: str) -> str:
    return f"reports/{file_name}"


def _save_report_content(file_name: str, content: bytes) -> tuple[str, str, int]:
    saved_name = default_storage.save(
        _report_storage_name(file_name),
        ContentFile(content, name=file_name),
    )
    return saved_name, file_name, len(content)


@shared_task(bind=True, max_retries=3)
def generate_report_run(self, report_run_id: str):
    """Generate ReportRun payload asynchronously."""
    from apps.reports.models import ReportRun, ReportRunStatus
    from apps.reports.services import ReportService

    report_run = ReportRun.objects.filter(id=report_run_id).first()
    if not report_run:
        logger.warning("ReportRun %s not found", report_run_id)
        return {"status": "not_found", "report_run_id": report_run_id}

    try:
        report_run.status = ReportRunStatus.RUNNING
        report_run.error_message = ""
        report_run.save(update_fields=["status", "error_message", "updated_at"])

        payload = ReportService.build_report_payload(
            report_type=report_run.report_type,
            parameters=report_run.parameters or {},
        )

        report_run.status = ReportRunStatus.COMPLETED
        report_run.result = payload
        report_run.save(update_fields=["status", "result", "updated_at"])

        return {"status": "ok", "report_run_id": str(report_run.id)}
    except Exception as exc:
        logger.exception("Failed generating ReportRun %s", report_run_id)
        report_run.status = ReportRunStatus.FAILED
        report_run.error_message = str(exc)
        report_run.save(update_fields=["status", "error_message", "updated_at"])
        raise self.retry(countdown=60 * (2**self.request.retries), exc=exc)


@shared_task(bind=True, max_retries=3)
def generate_report(self, report_request_id: int):
    """
    Generate report asynchronously.
    
    Steps:
    1. Set status to running
    2. Build dataset from filters
    3. Generate file (PDF/Excel/CSV)
    4. Save file path
    5. Set status to ready
    """
    from apps.reports.generators import REPORT_GENERATORS
    from apps.reports.models import ReportRequest, ReportStatus
    
    report = ReportRequest.objects.filter(id=report_request_id).first()
    if not report:
        logger.warning("ReportRequest %s not found", report_request_id)
        return
    
    try:
        # Set status to running
        report.status = ReportStatus.RUNNING
        report.save(update_fields=["status", "updated_at"])
        
        # Get generator
        generator = REPORT_GENERATORS.get(report.report_type)
        if not generator:
            raise ValueError(f"Unknown report type: {report.report_type}")
        
        # Build parameters
        filters = report.filters or {}
        kwargs = {}
        
        # Common filters
        if "date_from" in filters:
            kwargs["date_from"] = filters["date_from"]
        if "date_to" in filters:
            kwargs["date_to"] = filters["date_to"]
        if "member_id" in filters:
            kwargs["member_id"] = filters["member_id"]
        if "status" in filters:
            kwargs["status"] = filters["status"]
        if "entry_type" in filters:
            kwargs["entry_type"] = filters["entry_type"]
        
        # Generate data
        chama_id = report.chama_id if report.chama_id else 0
        data = generator(chama_id, **kwargs)
        
        # Generate file based on format
        if report.format == "csv":
            file_path, file_name, file_size = generate_csv(data, report.report_type)
        elif report.format == "xlsx":
            file_path, file_name, file_size = generate_excel(data, report.report_type)
        elif report.format == "pdf":
            file_path, file_name, file_size = generate_pdf(data, report.report_type)
        else:
            file_name = f"{report.report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            file_path, file_name, file_size = _save_report_content(
                file_name,
                json.dumps(data, indent=2, default=str).encode("utf-8"),
            )

        # Update report
        report.status = ReportStatus.READY
        report.file_path = file_path
        report.file_name = file_name
        report.file_size = file_size
        report.completed_at = timezone.now()
        report.save(update_fields=[
            "status", "file_path", "file_name", "file_size", 
            "completed_at", "updated_at"
        ])
        
        # Notify user
        send_report_ready_notification(report)
        
        logger.info("Report %s generated successfully: %s", report_request_id, file_name)
        
    except Exception as exc:
        logger.exception("Failed generating report %s", report_request_id)
        report.status = ReportStatus.FAILED
        report.error_message = str(exc)
        report.completed_at = timezone.now()
        report.save(update_fields=["status", "error_message", "completed_at", "updated_at"])
        
        raise self.retry(countdown=60 * (2**self.request.retries), exc=exc)


@shared_task
def process_scheduled_reports():
    """
    Check and process scheduled reports.
    Runs every minute via Celery Beat.
    """
    from apps.reports.models import ReportRequest, ScheduledReport
    
    now = timezone.now()
    
    # Find due scheduled reports
    due_reports = ScheduledReport.objects.filter(
        is_active=True,
        next_run_at__lte=now,
    )
    
    for scheduled in due_reports:
        # Create report request
        report_request = ReportRequest.objects.create(
            requested_by=scheduled.chama.created_by,  # Use chama creator
            chama=scheduled.chama,
            scope=scheduled.scope,
            report_type=scheduled.report_type,
            filters=scheduled.filters,
            format=scheduled.format,
            status="queued",
        )
        
        # Queue generation
        generate_report.delay(report_request.id)
        
        # Update scheduled report
        scheduled.last_run_at = now
        scheduled.last_status = "success"
        
        # Calculate next run (simplified - use croniter for full cron support)
        scheduled.next_run_at = now + timedelta(hours=24)  # Simplified
        scheduled.save()
        
        logger.info("Created scheduled report request: %s", report_request.id)


@shared_task
def cleanup_old_reports():
    """
    Clean up old report files.
    Run weekly.
    """
    from datetime import timedelta

    from apps.reports.models import ReportRequest
    
    # Delete reports older than 30 days
    cutoff = timezone.now() - timedelta(days=30)
    old_reports = ReportRequest.objects.filter(
        created_at__lt=cutoff,
        status__in=["ready", "failed"],
    )
    
    count = 0
    for report in old_reports:
        if report.file_path:
            try:
                default_storage.delete(report.file_path)
                count += 1
            except Exception:
                pass
    
    # Delete old records
    old_reports.delete()
    
    logger.info("Cleaned up %d old reports", count)


def generate_csv(data: dict, report_type: str) -> tuple:
    """Generate CSV file from report data."""
    import csv
    
    file_name = f"{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    output = io.StringIO()
    
    # Extract transactions for CSV
    transactions = data.get("transactions", [])
    if not transactions:
        # Try different keys
        for key in ["by_member", "by_month", "loans", "members"]:
            if key in data and isinstance(data[key], list):
                transactions = data[key]
                break
    
    if transactions:
        fieldnames = list(transactions[0].keys()) if transactions else []
        
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(transactions)
    
    return _save_report_content(file_name, output.getvalue().encode("utf-8"))


def generate_excel(data: dict, report_type: str) -> tuple:
    """Generate Excel file from report data."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        
        file_name = f"{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        wb = openpyxl.Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Add summary data
        row = 1
        if "summary" in data:
            for key, value in data["summary"].items():
                ws_summary.cell(row, 1, key)
                ws_summary.cell(row, 2, str(value))
                row += 1
        
        # Transactions sheet
        if "transactions" in data:
            ws_trans = wb.create_sheet("Transactions")
            transactions = data["transactions"]
            if transactions:
                headers = list(transactions[0].keys())
                for col, header in enumerate(headers, 1):
                    cell = ws_trans.cell(1, col, header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                for row_idx, tx in enumerate(transactions, 2):
                    for col_idx, header in enumerate(headers, 1):
                        ws_trans.cell(row_idx, col_idx, tx.get(header, ""))
        
        output = io.BytesIO()
        wb.save(output)
        return _save_report_content(file_name, output.getvalue())
        
    except ImportError:
        # Fallback to CSV
        return generate_csv(data, report_type)


def generate_pdf(data: dict, report_type: str) -> tuple:
    """Generate PDF file from report data."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
        
        file_name = f"{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title = data.get("chama", {}).get("name", "Report")
        elements.append(Paragraph(title, styles["Title"]))
        elements.append(Spacer(1, 12))
        
        # Period
        if "period" in data:
            period = data["period"]
            elements.append(Paragraph(f"Period: {period.get('from', '')} to {period.get('to', '')}", styles["Normal"]))
            elements.append(Spacer(1, 12))
        
        # Summary
        if "summary" in data:
            elements.append(Paragraph("Summary", styles["Heading2"]))
            summary_data = [["Metric", "Value"]]
            for key, value in data["summary"].items():
                summary_data.append([key.replace("_", " ").title(), str(value)])
            
            table = Table(summary_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(table)
        
        # Footer
        elements.append(Spacer(1, 30))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles["Normal"]))
        
        doc.build(elements)
        return _save_report_content(file_name, output.getvalue())
        
    except ImportError:
        # Fallback - just save JSON
        file_name = f"{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        return _save_report_content(
            file_name,
            json.dumps(data, indent=2, default=str).encode("utf-8"),
        )


def send_report_ready_notification(report):
    """Send notification when report is ready."""
    try:
        from apps.notifications.services import NotificationService
        
        NotificationService.send_push_notification(
            user=report.requested_by,
            title="Report Ready",
            body=f"Your {report.get_report_type_display()} report is ready for download.",
            data={"report_id": report.id, "type": "report_ready"}
        )
    except Exception:
        logger.exception("Failed to send notification for report %s", report.id)


# Import timedelta for scheduled reports
